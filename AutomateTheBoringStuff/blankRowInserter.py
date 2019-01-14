import sys
import openpyxl

def check_input():
    return len(sys.argv) == 4 and str.isdigit(sys.argv[1]) and int(sys.argv[1]) > 0 and str.isdigit(sys.argv[2]) and int(sys.argv[2]) > 0

def main():
    if not check_input(): sys.exit(1)
    
    N = int(sys.argv[1])
    M = int(sys.argv[2])
    
    wb = openpyxl.load_workbook(sys.argv[3])
    sheet = wb.active

    mc = sheet.max_column
    mr = sheet.max_row
    cell_value = [[None for i in range(mc)] for j in range(mr)]
    
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            cell_value[i-1][j-1] = sheet.cell(row=i, column=j).value

    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            sheet.cell(row=(i if i < N else i+M), column=j).value = cell_value[i-1][j-1]
            
    wb.save("updated_"+sys.argv[3])
        
if __name__ == '__main__':
    main()
