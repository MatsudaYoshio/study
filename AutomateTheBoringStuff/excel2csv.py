import os
import openpyxl
import csv

def main():
    for file_name in os.listdir('.'):
        if not file_name.endswith(".xlsx"): continue

        wb = openpyxl.load_workbook(file_name)
        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            csv_file = open(file_name[:-5]+"_"+sheet_name+".csv", 'w', newline='')
            csv_writer = csv.writer(csv_file)

            for i in range(1, sheet.max_row+1):
                row_data = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column+1)]

                csv_writer.writerow(row_data)

            csv_file.close()
        
if __name__ == '__main__':
    main()
