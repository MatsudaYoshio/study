import sys
import openpyxl
from openpyxl.styles import Font

def check_input():
    return len(sys.argv) == 2 and str.isdigit(sys.argv[1]) and int(sys.argv[1]) > 0

def main():
    if not check_input():
        print("Please input positive number!!")
        sys.exit(1)
    
    N = int(sys.argv[1])

    wb = openpyxl.Workbook()
    sheet = wb.active
    
    for i in range(1, N+1):
        sheet.cell(row=i+1, column=1).value = sheet.cell(row=1, column=i+1).value = i
        sheet.cell(row=i+1, column=1).font = sheet.cell(row=1, column=i+1).font = Font(bold=True)

    for i in range(2, N+2):
        for j in range(2, N+2):
            sheet.cell(row=i, column=j).value = sheet.cell(row=i, column=1).value*sheet.cell(row=1, column=j).value
        
    wb.save("multiplication_table.xlsx")
        
if __name__ == '__main__':
    main()
