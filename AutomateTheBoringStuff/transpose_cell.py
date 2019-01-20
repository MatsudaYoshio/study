import sys
import openpyxl

def main():
    if len(sys.argv) != 2:
        sys.exit(1)

    wb = openpyxl.load_workbook(sys.argv[1])
    sheet = wb.active
    
    mc = sheet.max_column
    mr = sheet.max_row
    cell_value = [[None for i in range(mc)] for j in range(mr)]
    
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            cell_value[i-1][j-1] = sheet.cell(row=i, column=j).value
            sheet.cell(row=i, column=j).value = 0
            
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            sheet.cell(row=i, column=j).value = cell_value[j-1][i-1]

    wb.save(sys.argv[1])
        
if __name__ == '__main__':
    main()
